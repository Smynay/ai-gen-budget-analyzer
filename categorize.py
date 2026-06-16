import os
import re
import yaml
from collections import defaultdict
from openpyxl import load_workbook

RESULTS_DIR = "results"
UNKNOWN = "Неизвестно"

# Prefixes to strip before normalization
PREFIXES = [
    r"ПРОЧИЕ ОПЕРАЦИИ TИНЬKОФФ\. ОПЕРАЦИЯ ПО СЧЕТУ \*{4}\d+ ",
    r"ПРОЧИЕ ОПЕРАЦИИ ТИНЬКОФФ\. ОПЕРАЦИЯ ПО СЧЕТУ \*{4}\d+ ",
    r"ПРОЧИЕ РАСХОДЫ ",
    r"ПРОЧИЕ ОПЕРАЦИИ ",
    r"ОПЛАТА УСЛУГ ",
    r"БЕЗНАЛИЧНАЯ ОПЛАТА ",
    r"СНЯТИЕ НАЛИЧНЫХ ",
    r"ВЫДАЧА НАЛИЧНЫХ\. ",
    r"ПОПОЛНЕНИЕ КАРТЫ ",
    r"ЗАЧИСЛЕНИЕ ЗАРАБОТНОЙ ПЛАТЫ ",
    r"ЗАЧИСЛЕНИЕ ",
    r"СПИСАНИЕ ",
]


def find_latest_xlsx():
    if not os.path.isdir(RESULTS_DIR):
        return None
    dirs = [d for d in os.listdir(RESULTS_DIR) if os.path.isdir(os.path.join(RESULTS_DIR, d))]
    try:
        dirs.sort(key=lambda d: (d[:10], int(d.split("_")[-1])))
    except ValueError:
        return None
    for d in reversed(dirs):
        xlsx = os.path.join(RESULTS_DIR, d, "excel.xlsx")
        if os.path.exists(xlsx):
            return xlsx
    return None


def normalize(desc):
    text = str(desc).upper().strip()
    for p in PREFIXES:
        text = re.sub(p, "", text)
    text = re.sub(r"НОМЕР КАРТЫ \*{4}\d+", "", text)
    text = re.sub(r"КАРТА \*{4}\d+", "", text)
    text = re.sub(r"\bСЧЕТУ \*{4}\d+\b", "СЧЕТУ ****", text)
    text = re.sub(r"\bСЧЕТ \d+\b", "СЧЕТ {N}", text)
    text = re.sub(r"\bДОГОВОР \d+\b", "ДОГОВОР {N}", text)
    text = re.sub(r"\bДОГОВОРА \d+\b", "ДОГОВОРА {N}", text)
    text = re.sub(r"\b(?<!\*{4})\d{4,}\b", "{N}", text)
    text = re.sub(r"\b\d{3,}\b", "{N}", text)
    text = re.sub(r"\s+", " ", text).strip(" .,;:!?")
    return text


def read_xlsx_categories(xlsx_path):
    wb = load_workbook(xlsx_path)
    ws = wb["Сводный"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {h: i for i, h in enumerate(headers)}
    cat_idx = col["Категория"]
    desc_idx = col["Описание"]
    amt_idx = col["Сумма"]

    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        rows.append({
            "row": r,
            "category": str(r[cat_idx]) if r[cat_idx] else UNKNOWN,
            "description": str(r[desc_idx] or ""),
            "amount": float(r[amt_idx]) if r[amt_idx] is not None else 0.0,
        })
    return wb, ws, col, rows


def get_unknown_groups(rows):
    groups = defaultdict(lambda: {
        "count": 0,
        "descriptions": [],
        "total_amount": 0.0,
        "row_indices": [],
    })
    for i, r in enumerate(rows):
        if r["category"] != UNKNOWN:
            continue
        key = normalize(r["description"])
        if not key:
            key = "(пусто)"
        g = groups[key]
        g["count"] += 1
        if r["description"] not in g["descriptions"]:
            g["descriptions"].append(r["description"])
        g["total_amount"] += r["amount"]
        g["row_indices"].append(i)
    return groups


def dump_mode(xlsx_path):
    _, _, _, rows = read_xlsx_categories(xlsx_path)
    groups = get_unknown_groups(rows)

    sorted_groups = sorted(
        groups.items(),
        key=lambda x: -x[1]["count"]
    )

    out = {"groups": []}
    for key, data in sorted_groups:
        out["groups"].append({
            "key": key,
            "count": data["count"],
            "descriptions": data["descriptions"],
            "total_amount": round(data["total_amount"], 2),
            "category": None,
        })

    out_path = os.path.join(os.path.dirname(xlsx_path), "descriptions.yaml")
    with open(out_path, "w", encoding="utf-8") as f:
        yaml.dump(out, f, allow_unicode=True, default_flow_style=False, sort_keys=False)
    print(f"Dumped {len(out['groups'])} groups to {out_path}")
    return out_path


def apply_mode(xlsx_path, yaml_path):
    with open(yaml_path, encoding="utf-8") as f:
        data = yaml.safe_load(f)

    category_map = {}
    for g in data["groups"]:
        cat = g.get("category")
        if cat and cat != UNKNOWN:
            for d in g["descriptions"]:
                category_map[normalize(d)] = cat

    if not category_map:
        print("No categories assigned in YAML")
        return

    wb = load_workbook(xlsx_path)

    for ws in wb.worksheets:
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        col = {h: i for i, h in enumerate(headers)}
        if "Категория" not in col or "Описание" not in col:
            continue
        cat_idx = col["Категория"] + 1
        desc_idx = col["Описание"] + 1

        updated = 0
        for r in ws.iter_rows(min_row=2, max_col=max(cat_idx, desc_idx) + 1):
            cell_cat = r[cat_idx - 1]
            cell_desc = r[desc_idx - 1]
            if cell_cat.value and str(cell_cat.value) != UNKNOWN:
                continue
            desc = str(cell_desc.value or "")
            key = normalize(desc)
            if key in category_map:
                cell_cat.value = category_map[key]
                updated += 1

        print(f"  Sheet '{ws.title}': updated {updated} rows")

    wb.save(xlsx_path)
    print(f"Saved: {xlsx_path}")


def main():
    import argparse
    parser = argparse.ArgumentParser(description="Categorize unknown transactions")
    parser.add_argument("--dump", action="store_true", help="Dump unknown transactions to YAML")
    parser.add_argument("--apply", metavar="YAML", nargs="?", const="auto", help="Apply categories from YAML to XLSX")
    args = parser.parse_args()

    xlsx_path = find_latest_xlsx()
    if not xlsx_path:
        print("No XLSX found in results/")
        return

    if args.dump:
        dump_mode(xlsx_path)
    elif args.apply:
        if args.apply == "auto":
            yaml_path = os.path.join(os.path.dirname(xlsx_path), "descriptions.yaml")
        else:
            yaml_path = args.apply
        if not os.path.exists(yaml_path):
            print(f"YAML not found: {yaml_path}")
            return
        apply_mode(xlsx_path, yaml_path)
    else:
        parser.print_help()


if __name__ == "__main__":
    main()
