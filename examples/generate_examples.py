import os, json, random
from datetime import datetime, timedelta
from collections import defaultdict
from openpyxl import Workbook

OUT_DIR = "examples"
os.makedirs(OUT_DIR, exist_ok=True)

MONTHS_RU = ["", "январь", "февраль", "март", "апрель", "май", "июнь",
              "июль", "август", "сентябрь", "октябрь", "ноябрь", "декабрь"]

SCHEMA = [
    {"key": "date", "name": "Дата"},
    {"key": "amount", "name": "Сумма"},
    {"key": "currency", "name": "Валюта"},
    {"key": "category", "name": "Категория"},
    {"key": "description", "name": "Описание"},
    {"key": "source", "name": "Источник"},
]

SOURCES = [
    ("Sberbank", [
        ("Доходы", "Зачисление заработной платы", 45000, 95000),
        ("Доходы", "Пополнение карты", 5000, 15000),
        ("Доходы", "Зачисление по договору", 20000, 40000),
    ]),
    ("Tinkoff", [
        ("Возврат долга", "Перевод на карту от А. ИВАНОВ. Номер карты ****0000", 1000, 5000),
        ("Возврат долга", "Перевод на карту от Е. ПЕТРОВ. Номер карты ****0000", 500, 3000),
    ]),
]

CATEGORIES = {
    "Супермаркеты": [
        "Безналичная оплата PYATEROCHKA 10136 MOSCOW RUS. Номер карты ****0000",
        "Безналичная оплата MAGNIT GM MOSCOW RUS. Номер карты ****0000",
        "Безналичная оплата LENTA-0010 SANKT-PETERBU RUS. Номер карты ****0000",
        "Безналичная оплата PEREKRESTOK MOSCOW RUS. Номер карты ****0000",
        "Безналичная оплата VKUSVILL 2224 MOSCOW RUS. Номер карты ****0000",
    ],
    "Рестораны и доставка еды": [
        "Безналичная оплата BURGER KING 0211 MOSCOW RUS. Номер карты ****0000",
        "Безналичная оплата YANDEX*5411*LAVKA MOSCOW RUS. Номер карты ****0000",
        "Безналичная оплата YANDEX*5814*EDA MOSCOW RUS. Номер карты ****0000",
        "Безналичная оплата KHLEBNIK VYBORG RUS. Номер карты ****0000",
        "Безналичная оплата COFFEESHOP COMPANY SANKT-PETERBU RUS. Номер карты ****0000",
    ],
    "Транспорт": [
        "Безналичная оплата TAXI YANDEX*4121*GO MOSCOW RUS. Номер карты ****0000",
        "Безналичная оплата METRO SANKT-PETERBU RUS. Номер карты ****0000",
        "Безналичная оплата WHOOSH MOSCOW RUS. Номер карты ****0000",
        "Безналичная оплата AZS ЗАПРАВКА MOSCOW RUS. Номер карты ****0000",
        "Безналичная оплата AVTOBUSNYJ PARK MOSCOW RUS. Номер карты ****0000",
    ],
    "Переводы": [
        "Перевод СБП по номеру телефона. Номер карты ****0000",
        "Перевод с карты на карту. Номер карты ****0000",
        "MAPP_SBERBANK_ONL@IN_PAY. Номер карты ****0000",
        "Перевод СБП на карту. Номер карты ****0000",
    ],
    "Онлайн-сервисы и подписки": [
        "Безналичная оплата YANDEX*5815*PLUS MOSCOW RUS. Номер карты ****0000",
        "Безналичная оплата WILDBERRIES SBERPAY MOSCOW RUS. Номер карты ****0000",
        "Безналичная оплата OZON MOSCOW RUS. Номер карты ****0000",
        "Безналичная оплата STEAM GAMES MOSCOW RUS. Номер карты ****0000",
        "Безналичная оплата GOOGLE SERVICES MOSCOW RUS. Номер карты ****0000",
    ],
    "Здоровье и красота": [
        "Безналичная оплата АПТЕКА 24 MOSCOW RUS. Номер карты ****0000",
        "Безналичная оплата КЛИНИКА ЗДОРОВЬЕ MOSCOW RUS. Номер карты ****0000",
        "Безналичная оплата ВЕТЕРИНАРНАЯ КЛИНИКА MOSCOW RUS. Номер карты ****0000",
    ],
    "Коммунальные и связь": [
        "Оплата услуг СВЯЗЬ ИНТЕРНЕТ. Номер карты ****0000",
        "Коммунальные платежи. Номер карты ****0000",
        "SBERCHAEVYE КОММУНАЛЬНЫЕ. Номер карты ****0000",
    ],
    "Товары для дома": [
        "Покупка OLALA VIBORG RUS. Номер карты ****0000",
        "Покупка ULYBKA MOSCOW RUS. Номер карты ****0000",
        "Товары для дома VSEM DLYA DOMA. Номер карты ****0000",
    ],
    "Развлечения": [
        "Билеты КИНОТЕАТР MOSCOW RUS. Номер карты ****0000",
        "Оплата ОТДЫХ И РАЗВЛЕЧЕНИЯ. Номер карты ****0000",
        "Безналичная оплата ТЕАТР MOSCOW RUS. Номер карты ****0000",
    ],
    "Наличные": [
        "Снятие наличных ATM 123456. Номер карты ****0000",
        "Выдача наличных. Номер карты ****0000",
        "СНЯТИЕ НАЛИЧНЫХ. Номер карты ****0000",
    ],
    "Неизвестно": [
        "Оплата услуг AVITO MOSCOW RUS. Номер карты ****0000",
        "Безналичная оплата NOTARY SERVICES MOSCOW RUS. Номер карты ****0000",
        "Безналичная оплата BINBRAIN LLC SANKT-PETERBU RUS. Номер карты ****0000",
        "Прочие операции. Номер карты ****0000",
    ],
}


def random_date(year=2026):
    start = datetime(year, 1, 1)
    end = datetime(year, 12, 31, 23, 59)
    delta = (end - start).total_seconds()
    return start + timedelta(seconds=random.random() * delta)


def round_amount(v):
    return round(v, 2)


def generate_txns(count=100):
    txns = []
    all_cats = list(CATEGORIES.items())
    n_per_cat = max(2, count // (len(all_cats) + 2))  # +2 for Доходы and Возврат долга

    for cat, descs in all_cats:
        for _ in range(n_per_cat):
            dt = random_date()
            desc = random.choice(descs)
            amt = -round_amount(random.uniform(100, 15000))
            source = random.choice(["Sberbank", "Tinkoff"])
            txns.append({
                "date": dt,
                "amount": amt,
                "currency": "RUB",
                "category": cat,
                "description": desc,
                "source": f"{source}: Vypiska.pdf",
            })

    income_templates = [
        ("Доходы", 45000, 95000, "Зачисление заработной платы"),
        ("Доходы", 5000, 15000, "Пополнение карты"),
        ("Доходы", 20000, 40000, "Зачисление по договору"),
        ("Возврат долга", 500, 5000, "Перевод на карту от А. ИВАНОВ. Номер карты ****0000"),
        ("Возврат долга", 300, 3000, "Перевод на карту от Е. ПЕТРОВА. Номер карты ****0000"),
    ]
    for cat, lo, hi, desc in income_templates:
        for _ in range(n_per_cat // 2 + 1):
            dt = random_date()
            amt = round_amount(random.uniform(lo, hi))
            source = random.choice(["Sberbank", "Tinkoff"])
            txns.append({
                "date": dt,
                "amount": amt,
                "currency": "RUB",
                "category": cat,
                "description": desc,
                "source": f"{source}: Vypiska.pdf",
            })

    # Shuffle and trim to exact count
    random.shuffle(txns)
    txns = txns[:count]
    while len(txns) < count:
        dt = random_date()
        cat = random.choice(list(CATEGORIES.keys()))
        desc = random.choice(CATEGORIES[cat])
        amt = -round_amount(random.uniform(100, 15000))
        source = random.choice(["Sberbank", "Tinkoff"])
        txns.append({
            "date": dt, "amount": amt, "currency": "RUB",
            "category": cat, "description": desc,
            "source": f"{source}: Vypiska.pdf",
        })
    txns.sort(key=lambda t: t["date"])
    return txns


def write_xlsx(txns):
    col_keys = [c["key"] for c in SCHEMA]
    col_names = [c["name"] for c in SCHEMA]

    wb = Workbook()
    ws = wb.active
    ws.title = "Сводный"
    ws.append(col_names)

    # Per-source sheets
    sources = defaultdict(list)
    for t in txns:
        src_key = "Sberbank" if "Sberbank" in t["source"] else "Tinkoff"
        sources[src_key].append(t)

    source_sheets = {}
    for src_name in sources:
        ws_src = wb.create_sheet(title=src_name)
        no_source_names = [c["name"] for c in SCHEMA if c["key"] != "source"]
        ws_src.append(no_source_names)
        source_sheets[src_name] = ws_src

    for t in txns:
        ws.append([t.get(k, "") for k in col_keys])
        src_key = "Sberbank" if "Sberbank" in t["source"] else "Tinkoff"
        no_source_keys = [k for k in col_keys if k != "source"]
        source_sheets[src_key].append([t.get(k, "") for k in no_source_keys])

    date_col_idx = col_keys.index("date") + 1
    for ws_to_fmt in [ws] + list(source_sheets.values()):
        for row in ws_to_fmt.iter_rows(min_row=2, min_col=date_col_idx, max_col=date_col_idx):
            for cell in row:
                if isinstance(cell.value, datetime):
                    cell.number_format = "DD.MM.YYYY HH:MM"

    xlsx_path = os.path.join(OUT_DIR, "excel.xlsx")
    wb.save(xlsx_path)
    return xlsx_path


HTML_TPL = r"""<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Финансовый отчёт</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.7/dist/chart.umd.min.js"></script>
<style>
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,sans-serif;background:#f0f2f5;color:#222;max-width:1100px;margin:0 auto;padding:24px 16px}
h1{font-size:24px;margin-bottom:20px;color:#111}
.summary{display:grid;grid-template-columns:repeat(4,1fr);gap:12px;margin-bottom:24px}
.card{background:#fff;border-radius:10px;padding:16px;box-shadow:0 1px 3px rgba(0,0,0,.08)}
.card h3{font-size:11px;color:#888;text-transform:uppercase;letter-spacing:.5px;margin-bottom:4px}
.card .val{font-size:22px;font-weight:700}
.card .val.neg{color:#e74c3c}
.card .val.pos{color:#27ae60}
.tabs{display:flex;gap:8px;margin-bottom:20px}
.tab{padding:8px 20px;border:1px solid #ddd;border-radius:20px;background:#fff;cursor:pointer;font-size:14px;transition:.15s;user-select:none}
.tab:hover{border-color:#999}
.tab.active{background:#333;color:#fff;border-color:#333}
.chart-row{display:flex;gap:20px;align-items:flex-start}
.chart-box{flex:1;background:#fff;border-radius:10px;padding:20px;box-shadow:0 1px 3px rgba(0,0,0,.08);max-width:500px}
.cat-table{flex:1;background:#fff;border-radius:10px;padding:20px;box-shadow:0 1px 3px rgba(0,0,0,.08)}
table{width:100%;border-collapse:collapse;font-size:14px}
th{text-align:left;padding:8px 8px 8px 0;color:#888;font-weight:600;border-bottom:1px solid #eee;font-size:12px;text-transform:uppercase;letter-spacing:.3px}
td{padding:8px 8px 8px 0;border-bottom:1px solid #f5f5f5;cursor:pointer}
tr:last-child td{border-bottom:none}
tr:hover td{background:#f5f5f5}
.cat-dot{display:inline-block;width:10px;height:10px;border-radius:50%;margin-right:8px;vertical-align:middle}
.amt{text-align:right;font-variant-numeric:tabular-nums;white-space:nowrap}
.amt.neg{color:#e74c3c}
.amt.pos{color:#27ae60}
#txnSection{display:none;margin-top:24px;background:#fff;border-radius:10px;padding:20px;box-shadow:0 1px 3px rgba(0,0,0,.08)}
#txnSection h2{font-size:16px;margin-bottom:12px}
#txnSection h2 .close{float:right;cursor:pointer;color:#aaa;font-size:22px;line-height:1}
#txnSection h2 .close:hover{color:#333}
#txnTable{width:100%}
#txnTable td{border-bottom:1px solid #eee;padding:6px 8px 6px 0;cursor:default}
#txnTable th:nth-child(1){width:140px}
#txnTable th:nth-child(2){width:130px;text-align:right}
.hide-source #txnTable th:nth-child(4),.hide-source #txnTable td:nth-child(4){display:none}
.toggle-source{font-size:13px;margin-bottom:12px;display:inline-block;cursor:pointer;user-select:none}
.toggle-source input{vertical-align:middle;margin-right:4px}
</style>
</head>
<body>
<h1>&#x1F4CA; $REPORT_TITLE</h1>
<div class="summary" id="summaryCards"></div>
<div class="tabs" id="tabs">
    <div class="tab active" data-mode="expense">Расходы</div>
    <div class="tab" data-mode="income">Доходы</div>
    <div class="tab" data-mode="all">Всё</div>
</div>
<div class="chart-row">
    <div class="chart-box"><canvas id="pieChart"></canvas></div>
    <div class="cat-table" id="catTableContainer"></div>
</div>
<div id="txnSection">
    <h2><span id="txnTitle"></span><span class="close" onclick="closeTxn()">&times;</span></h2>
    <label class="toggle-source"><input type="checkbox" id="showSource" onchange="toggleSource()"> Показывать источник</label>
    <div id="txnWrapper">
    <table id="txnTable"><thead><tr><th onclick="sortBy('d')" style="cursor:pointer">Дата <span id="sortInd"></span></th><th>Сумма</th><th>Описание</th><th>Источник</th></tr></thead><tbody id="txnBody"></tbody></table>
    </div>
</div>
<script>
var DATA = $JSON_DATA ;

var COLORS = [
    '#e74c3c','#3498db','#2ecc71','#f39c12','#9b59b6','#1abc9c','#e67e22','#34495e',
    '#c0392b','#2980b9','#27ae60','#d35400','#8e44ad','#16a085','#f1c40f','#2c3e50',
    '#e88','#698','#8a9','#ba8','#9af','#fa8','#a78','#89a'
];

var currentMode = 'expense';
var chart = null;
var categoryColors = {};
var sortCol = 'd';
var sortDir = -1;

(function() {
    var i = 0;
    for (var name in DATA.categories) {
        categoryColors[name] = COLORS[i % COLORS.length];
        i++;
    }
})();

function filterData(mode) {
    var cats = {};
    for (var name in DATA.categories) {
        var d = DATA.categories[name];
        if (mode === 'expense' && d.total >= 0) continue;
        if (mode === 'income' && d.total <= 0) continue;
        cats[name] = d;
    }
    return cats;
}

function fmt(v) {
    var neg = v < 0;
    var s = Math.abs(v).toFixed(2);
    s = s.replace(/\B(?=(\d{3})+(?!\d))/g, ' ');
    return (neg ? '\u2212' : '') + s;
}

function fmtDate(s) {
    var p = s.split(' ');
    var d = p[0].split('-');
    return d[2] + '.' + d[1] + '.' + d[0] + ' ' + (p[1] || '');
}

function renderChart(mode) {
    var cats = filterData(mode);
    if (chart) chart.destroy();
    var ctx = document.getElementById('pieChart').getContext('2d');
    chart = new Chart(ctx, {
        type: 'pie',
        data: { labels: Object.keys(cats), datasets: [{ data: Object.values(cats).map(function(c) { return Math.abs(c.total); }), backgroundColor: Object.keys(cats).map(function(k) { return categoryColors[k] || '#ccc'; }), borderWidth: 0 }] },
        options: {
            responsive: true,
            plugins: {
                legend: { display: false },
                tooltip: {
                    callbacks: {
                        label: function(ctx) {
                            var total = ctx.dataset.data.reduce(function(a,b){return a+b}, 0);
                            var pct = ((ctx.parsed / total) * 100).toFixed(1);
                            var cats = filterData(currentMode);
                            return ' ' + ctx.label + ': ' + fmt(cats[ctx.label].total) + ' \u20bd (' + pct + '%)';
                        }
                    }
                }
            },
            onClick: function(e, elements) {
                if (elements.length > 0) {
                    showTransactions(this.data.labels[elements[0].index]);
                }
            }
        }
    });
    renderCategoryTable(cats);
}

function renderCategoryTable(cats) {
    var html = '<table><thead><tr><th>Категория</th><th class="amt">Сумма</th><th class="amt">Кол-во</th></tr></thead><tbody>';
    var keys = Object.keys(cats);
    for (var i = 0; i < keys.length; i++) {
        var name = keys[i];
        var d = cats[name];
        var color = categoryColors[name] || '#ccc';
        html += '<tr onclick="showTransactions(\'' + name.replace(/'/g, "\\'") + '\')">'
            + '<td><span class="cat-dot" style="background:' + color + '"></span>' + name + '</td>'
            + '<td class="amt ' + (d.total < 0 ? 'neg' : 'pos') + '">' + fmt(d.total) + '</td>'
            + '<td class="amt">' + d.count + '</td></tr>';
    }
    html += '</tbody></table>';
    document.getElementById('catTableContainer').innerHTML = html;
}

function doSort(txns) {
    txns.sort(function(a, b) {
        var va = a[sortCol], vb = b[sortCol];
        if (sortCol === 'd') return (va < vb ? -1 : va > vb ? 1 : 0) * sortDir;
        if (sortCol === 'a') return (parseFloat(va) - parseFloat(vb)) * sortDir;
        return (va < vb ? -1 : va > vb ? 1 : 0) * sortDir;
    });
}

function sortBy(col) {
    if (col === sortCol) sortDir *= -1;
    else { sortCol = col; sortDir = -1; }
    var cur = document.getElementById('txnTitle').textContent;
    if (cur) showTransactions(cur);
}

function showTransactions(catName) {
    var d = DATA.categories[catName];
    if (!d) return;
    document.getElementById('txnTitle').textContent = catName;
    var sorted = d.transactions.slice();
    doSort(sorted);
    var html = '';
    for (var i = 0; i < sorted.length; i++) {
        var t = sorted[i];
        var amt = parseFloat(t.a);
        var cls = amt < 0 ? 'neg' : 'pos';
        html += '<tr><td>' + fmtDate(t.d) + '</td><td class="amt ' + cls + '">' + fmt(amt) + '</td><td>' + t.n.slice(0,120) + '</td><td>' + t.s + '</td></tr>';
    }
    document.getElementById('txnBody').innerHTML = html;
    document.getElementById('sortInd').textContent = sortCol === 'd' ? (sortDir < 0 ? ' \u2193' : ' \u2191') : '';
    document.getElementById('txnSection').style.display = 'block';
}

function closeTxn() {
    document.getElementById('txnSection').style.display = 'none';
}

function toggleSource() {
    var el = document.getElementById('txnWrapper');
    if (document.getElementById('showSource').checked) {
        el.classList.remove('hide-source');
    } else {
        el.classList.add('hide-source');
    }
}
toggleSource();

var s = DATA.summary;
document.getElementById('summaryCards').innerHTML =
    '<div class="card"><h3>Транзакции</h3><div class="val">' + s.count + '</div></div>'
    + '<div class="card"><h3>Доходы</h3><div class="val pos">' + fmt(s.income) + '</div></div>'
    + '<div class="card"><h3>Расходы</h3><div class="val neg">' + fmt(s.expense) + '</div></div>'
    + '<div class="card"><h3>Нетто</h3><div class="val ' + (s.net < 0 ? 'neg' : 'pos') + '">' + fmt(s.net) + '</div></div>';

renderChart('expense');

(function() {
    var tabs = document.querySelectorAll('.tab');
    for (var i = 0; i < tabs.length; i++) {
        tabs[i].addEventListener('click', function() {
            var active = document.querySelectorAll('.tab');
            for (var j = 0; j < active.length; j++) active[j].classList.remove('active');
            this.classList.add('active');
            currentMode = this.dataset.mode;
            closeTxn();
            renderChart(currentMode);
        });
    }
})();
</script>
</body>
</html>"""


def build_report(txns):
    totals = {"count": 0, "income": 0.0, "expense": 0.0}
    categories = defaultdict(lambda: {"total": 0.0, "count": 0, "transactions": []})

    for t in txns:
        cat = t["category"]
        amount = t["amount"]
        date_str = t["date"].strftime("%Y-%m-%d %H:%M")

        categories[cat]["total"] += amount
        categories[cat]["count"] += 1
        categories[cat]["transactions"].append({
            "d": date_str, "a": round(amount, 2),
            "n": t["description"], "s": t["source"]
        })

        if amount > 0:
            totals["income"] += amount
        else:
            totals["expense"] += amount
        totals["count"] += 1

    data = {
        "summary": {
            "count": totals["count"],
            "income": round(totals["income"], 2),
            "expense": round(totals["expense"], 2),
            "net": round(totals["income"] + totals["expense"], 2),
        },
        "categories": {}
    }
    for cat_name in sorted(categories, key=lambda c: -abs(categories[c]["total"])):
        cd = categories[cat_name]
        cd["total"] = round(cd["total"], 2)
        cd["transactions"].sort(key=lambda t: t["d"])
        data["categories"][cat_name] = cd

    json_data = json.dumps(data, ensure_ascii=False, indent=2)
    html = HTML_TPL.replace("$JSON_DATA", json_data).replace("$REPORT_TITLE", "Финансовый отчёт (пример)")

    html_path = os.path.join(OUT_DIR, "index.html")
    with open(html_path, "w", encoding="utf-8") as f:
        f.write(html)
    return html_path


def main():
    random.seed(42)
    txns = generate_txns(100)
    xlsx_path = write_xlsx(txns)
    report_path = build_report(txns)
    print(f"Generated: {xlsx_path}")
    print(f"Generated: {report_path}")
    print(f"Total: {len(txns)} transactions, {len(CATEGORIES)} categories + Доходы + Возврат долга")


if __name__ == "__main__":
    main()
