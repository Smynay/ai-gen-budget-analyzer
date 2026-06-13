import os
import re
import yaml
from datetime import datetime
from collections import defaultdict
import fitz
from openpyxl import Workbook

CONFIG_PATH = "config.yaml"
SOURCES_DIR = "sources"
RESULTS_DIR = "results"

DATE_RE = re.compile(r"\d{2}\.\d{2}\.\d{4}")
DATETIME_RE = re.compile(r"(\d{2}\.\d{2}\.\d{4})\s+(\d{2}):(\d{2})")


def load_config(path):
    with open(path, encoding="utf-8") as f:
        return yaml.safe_load(f)


def strip_currency(text):
    return re.sub(r"[^\d,.\s+\-]", "", text).strip()


def parse_russian_amount(text, decimal_sep, group_sep):
    s = strip_currency(text)
    if not s:
        return None
    sign = -1.0 if s[0] == "-" else 1.0
    if s[0] in ("-", "+"):
        s = s[1:]
    s = s.replace(group_sep, "").replace(decimal_sep, ".")
    try:
        return sign * float(s)
    except ValueError:
        return None


def parse_russian_date(text):
    m = DATE_RE.search(text)
    if m:
        try:
            return datetime.strptime(m.group(0), "%d.%m.%Y")
        except ValueError:
            pass
    return None


def parse_russian_datetime(text):
    if not text:
        return None
    m = DATETIME_RE.search(text)
    if m:
        try:
            return datetime(
                int(m.group(1)[6:]), int(m.group(1)[3:5]), int(m.group(1)[:2]),
                int(m.group(2)), int(m.group(3))
            )
        except ValueError:
            pass
    return parse_russian_date(text)


def word_in_col(word, x_start, x_end):
    cx = (word[0] + word[2]) / 2
    return x_start <= cx <= x_end


def column_text(words, x_start, x_end):
    parts = [w[4] for w in words if word_in_col(w, x_start, x_end)]
    return " ".join(parts).strip()


def is_numeric_value(text):
    s = strip_currency(text)
    if not s:
        return False
    s = s.lstrip("+-")
    if not s:
        return False
    return bool(re.match(r"^[\d\s.,]+$", s))


def match_template(templates, doc):
    meta = doc.metadata
    meta_text = " ".join(v for v in meta.values() if v)
    first_page_text = doc[0].get_text() if len(doc) > 0 else ""
    haystack = (meta_text + " " + first_page_text).lower()
    for tmpl in templates:
        for kw in tmpl["match"]["keywords"]:
            if kw.lower() in haystack:
                return tmpl
    return None


def group_words_by_row(words, tol=3):
    rows = defaultdict(list)
    for w in words:
        y_key = round(w[1] / tol) * tol
        rows[y_key].append(w)
    sorted_rows = []
    for y in sorted(rows.keys()):
        sorted_rows.append((y, sorted(rows[y], key=lambda w: w[0])))
    return sorted_rows


def find_header_row(rows, all_cols, header_text):
    """Find first row whose text (any column) contains header_text."""
    for idx, (y, words) in enumerate(rows):
        full_text = " ".join(w[4] for w in words)
        if header_text.lower() in full_text.lower():
            return idx
    return -1


def extract_tinkoff(pdf_path, tmpl):
    cfg_cols = tmpl["columns"]
    amt_cfg = tmpl["amount"]
    doc = fitz.open(pdf_path)
    transactions = []
    started = False

    for page in doc:
        words = page.get_text("words")
        if not words:
            continue
        rows = group_words_by_row(words)

        if not started:
            hdr_idx = find_header_row(rows, cfg_cols, "Описание")
            if hdr_idx < 0:
                continue
            rows = rows[hdr_idx + 1 :]
            started = True

        txn = None
        desc_lines = []

        for y, row_words in rows:
            amt_text = column_text(row_words, cfg_cols["amount_txn"]["x_start"], cfg_cols["amount_txn"]["x_end"])
            has_amount = is_numeric_value(amt_text)

            if has_amount:
                if txn:
                    merged_desc = (txn.get("description", "") + " " + " ".join(desc_lines)).strip()
                    if merged_desc:
                        txn["description"] = merged_desc
                    transactions.append(txn)

                date_text = column_text(row_words, cfg_cols["date_op"]["x_start"], cfg_cols["date_op"]["x_end"])
                parsed_date = parse_russian_date(date_text)
                amount = parse_russian_amount(amt_text, amt_cfg["decimal_separator"], amt_cfg["group_separator"])
                desc = column_text(row_words, cfg_cols["description"]["x_start"], cfg_cols["description"]["x_end"])

                txn = {"date": parsed_date, "amount": amount, "description": desc}
                desc_lines = []
            else:
                if txn and len(desc_lines) < 4:
                    time_text = column_text(row_words, cfg_cols["date_op"]["x_start"], cfg_cols["date_op"]["x_end"])
                    tm = re.match(r"^(\d{2}):(\d{2})$", time_text.strip())
                    if tm:
                        date_val = txn.get("date")
                        if isinstance(date_val, datetime):
                            txn["date"] = date_val.replace(hour=int(tm.group(1)), minute=int(tm.group(2)))
                    desc_text = column_text(row_words, cfg_cols["description"]["x_start"], cfg_cols["description"]["x_end"])
                    if desc_text:
                        desc_lines.append(desc_text)

        if txn:
            merged_desc = (txn.get("description", "") + " " + " ".join(desc_lines)).strip()
            if merged_desc:
                txn["description"] = merged_desc
            transactions.append(txn)

    doc.close()

    merged = []
    for t in transactions:
        if t["date"] is not None and t["amount"] is not None:
            if t["description"]:
                parts = t["description"].split()
                seen = set()
                deduped = []
                for p in parts:
                    if p not in seen:
                        seen.add(p)
                        deduped.append(p)
                t["description"] = " ".join(deduped)
            merged.append(t)
    return merged


def extract_sberbank(pdf_path, tmpl):
    cfg_cols = tmpl["columns"]
    amt_cfg = tmpl["amount"]
    doc = fitz.open(pdf_path)
    transactions = []
    in_table = False

    for page in doc:
        words = page.get_text("words")
        if not words:
            continue
        rows = group_words_by_row(words)

        if not in_table:
            text = page.get_text()
            if "Расшифровка операций" not in text:
                continue
            hdr_idx = find_header_row(rows, cfg_cols, "ДАТА ОПЕРАЦИИ")
            if hdr_idx < 0:
                in_table = True
                continue
            rows = rows[hdr_idx + 1 :]
            in_table = True

        txn = None
        desc_lines = []

        for y, row_words in rows:
            date_text = column_text(row_words, cfg_cols["date"]["x_start"], cfg_cols["date"]["x_end"])
            cat_text = column_text(row_words, cfg_cols["category"]["x_start"], cfg_cols["category"]["x_end"])
            amt_text = column_text(row_words, cfg_cols["amount"]["x_start"], cfg_cols["amount"]["x_end"])

            has_amount = is_numeric_value(amt_text)
            has_date = bool(DATE_RE.search(date_text))

            if has_amount and has_date:
                if txn:
                    full_desc = (txn.get("description", "") + " " + " ".join(desc_lines)).strip()
                    if full_desc:
                        txn["description"] = full_desc
                    if txn["date"] is not None and txn["amount"] is not None:
                        transactions.append(txn)

                parsed_date = parse_russian_datetime(date_text)
                raw_amount = parse_russian_amount(amt_text, amt_cfg["decimal_separator"], amt_cfg["group_separator"])
                # Sberbank: unsigned amount = expense, signed with "+" = income
                if raw_amount is not None and not amt_text.strip().startswith("+"):
                    raw_amount = -abs(raw_amount)

                txn = {"date": parsed_date, "amount": raw_amount, "description": cat_text}
                desc_lines = []

            elif txn and len(desc_lines) < 4:
                if cat_text and any(kw in cat_text.upper() for kw in ("ПРОДОЛЖЕНИЕ", "СЛЕДУЮЩЕЙ", "СТРАНИЦЕ")):
                    continue
                if cat_text:
                    desc_lines.append(cat_text)
                elif date_text and not DATE_RE.match(date_text.strip()):
                    desc_lines.append(date_text)

        if txn:
            full_desc = (txn.get("description", "") + " " + " ".join(desc_lines)).strip()
            if full_desc:
                txn["description"] = full_desc
            if txn["date"] is not None and txn["amount"] is not None:
                transactions.append(txn)

    doc.close()

    merged = []
    for t in transactions:
        if t["date"] is not None and t["amount"] is not None:
            if t["description"]:
                parts = t["description"].split()
                seen = set()
                deduped = []
                for p in parts:
                    if p not in seen:
                        seen.add(p)
                        deduped.append(p)
                t["description"] = " ".join(deduped)
            merged.append(t)

    return merged


def categorize(description, amount, categories_cfg):
    if not categories_cfg or "items" not in categories_cfg:
        return categories_cfg.get("fallback", "Другое") if categories_cfg else "Другое"
    if not description:
        return categories_cfg.get("fallback", "Другое")
    text = str(description).upper()
    for item in categories_cfg["items"]:
        direction = item.get("direction")
        if direction == "income" and (amount is None or amount <= 0):
            continue
        if direction == "expense" and (amount is None or amount >= 0):
            continue
        for kw in item.get("keywords", []):
            if kw.upper() in text:
                return item["name"]
    return categories_cfg.get("fallback", "Другое")


def extract_transactions(pdf_path, template):
    name = template["name"]
    if name == "sberbank_debit":
        return extract_sberbank(pdf_path, template)
    elif name == "tinkoff_movement":
        return extract_tinkoff(pdf_path, template)
    raise ValueError(f"Unknown template: {name}")


def sheet_name_from_file(filename):
    name = os.path.splitext(filename)[0]
    name = re.sub(r"[^\w\s]", "_", name)
    if len(name) > 31:
        name = name[:28] + "..."
    return name


def write_xlsx(all_data, config, output_path):
    schema = config["unified_schema"]["columns"]
    col_keys = [c["key"] for c in schema]
    col_names = [c["name"] for c in schema]

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Сводный"
    ws_summary.append(col_names)

    source_sheets = {}
    for file_key in all_data:
        sheet_name = sheet_name_from_file(file_key)
        ws = wb.create_sheet(title=sheet_name)
        no_source_keys = [k for k in col_keys if k != "source"]
        no_source_names = [c["name"] for c in schema if c["key"] != "source"]
        ws.append(no_source_names)
        source_sheets[file_key] = ws

    for file_key, txns in all_data.items():
        ws = source_sheets[file_key]
        for t in txns:
            no_source_row = [t.get(k, "") for k in no_source_keys]
            ws.append(no_source_row)

            summary_row = [t.get(k, "") for k in col_keys]
            ws_summary.append(summary_row)

    date_col_idx = col_keys.index("date") + 1
    for ws in [ws_summary] + list(source_sheets.values()):
        for row in ws.iter_rows(min_row=2, min_col=date_col_idx, max_col=date_col_idx):
            for cell in row:
                if isinstance(cell.value, datetime):
                    cell.number_format = "DD.MM.YYYY HH:MM"

    wb.save(output_path)


def main():
    config = load_config(CONFIG_PATH)
    os.makedirs(RESULTS_DIR, exist_ok=True)

    ext = config.get("sources", {}).get("directory", SOURCES_DIR)
    pdf_files = sorted([f for f in os.listdir(ext) if f.lower().endswith(".pdf")])

    if not pdf_files:
        print("No PDF files found in sources/")
        return

    all_data = {}
    default_currency = "RUB"
    for c in config["unified_schema"]["columns"]:
        if c["key"] == "currency" and "default" in c:
            default_currency = c["default"]

    for fname in pdf_files:
        fpath = os.path.join(ext, fname)
        print(f"Processing: {fname}...")

        doc = fitz.open(fpath)
        template = match_template(config["templates"], doc)
        doc.close()

        if not template:
            print(f"  No matching template for {fname}, skipping")
            continue

        print(f"  Matched template: {template['name']}")

        txns = extract_transactions(fpath, template)

        categories_cfg = None
        personal_path = "config_personal.yaml"
        if os.path.exists(personal_path):
            personal_cfg = load_config(personal_path)
            categories_cfg = personal_cfg.get("categories")
            print("  Using personal categories from config_personal.yaml")
        if categories_cfg is None:
            categories_cfg = config.get("categories")
        currency = template.get("currency", default_currency)
        for t in txns:
            t["currency"] = currency
            t["source"] = fname
            t["category"] = categorize(t.get("description", ""), t.get("amount"), categories_cfg)

        txns.sort(key=lambda t: t["date"] if t["date"] else datetime.min)

        all_data[fname] = txns
        print(f"  Extracted {len(txns)} transactions")

    if not all_data:
        print("No transactions extracted.")
        return

    base = datetime.now().strftime("%Y-%m-%d")
    suffix = 1
    while True:
        dir_name = f"{base}_{suffix}"
        out_dir = os.path.join(RESULTS_DIR, dir_name)
        if not os.path.exists(out_dir):
            break
        suffix += 1
    os.makedirs(out_dir)

    xlsx_path = os.path.join(out_dir, "excel.xlsx")
    write_xlsx(all_data, config, xlsx_path)
    print(f"Saved: {out_dir}/")

    # clean up old flat xlsx files
    for f in os.listdir(RESULTS_DIR):
        fpath = os.path.join(RESULTS_DIR, f)
        if os.path.isfile(fpath) and f.lower().endswith(".xlsx"):
            os.remove(fpath)


if __name__ == "__main__":
    main()
