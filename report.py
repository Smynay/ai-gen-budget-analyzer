import os
import re
import sys
import json
from collections import defaultdict
from datetime import datetime
from openpyxl import load_workbook

RESULTS_DIR = "results"

MONTHS_RU = ["", "январь", "февраль", "март", "апрель", "май", "июнь",
              "июль", "август", "сентябрь", "октябрь", "ноябрь", "декабрь"]

HTML_TPL = """<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Финансовый отчёт</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.7/dist/chart.umd.min.js"></script>
<style>
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,sans-serif;background:#f0f2f5;color:#222;max-width:1100px;margin:0 auto;padding:24px 16px}
h1{font-size:24px;margin-bottom:20px;color:#111}
.summary{display:grid;grid-template-columns:repeat(4,1fr);gap:12px;margin-bottom:24px}
.card{background:#fff;border-radius:10px;padding:16px;box-shadow:0 1px 3px rgba(0,0,0,.08)}
.card h3{font-size:11px;color:#888;text-transform:uppercase;letter-spacing:.5px;margin-bottom:4px}
.card .val{font-size:22px;font-weight:700}
.card .val.neg{color:#e74c3c}
.card .val.pos{color:#27ae60}
.tabs{display:flex;gap:8px;margin-bottom:20px}
.tab{padding:8px 20px;border:1px solid #ddd;border-radius:20px;background:#fff;cursor:pointer;font-size:14px;transition:.15s;user-select:none}
.tab:hover{border-color:#999}
.tab.active{background:#333;color:#fff;border-color:#333}
.chart-row{display:flex;gap:20px;align-items:flex-start}
.chart-box{flex:1;background:#fff;border-radius:10px;padding:20px;box-shadow:0 1px 3px rgba(0,0,0,.08);max-width:500px}
.cat-table{flex:1;background:#fff;border-radius:10px;padding:20px;box-shadow:0 1px 3px rgba(0,0,0,.08)}
table{width:100%;border-collapse:collapse;font-size:14px}
th{text-align:left;padding:8px 8px 8px 0;color:#888;font-weight:600;border-bottom:1px solid #eee;font-size:12px;text-transform:uppercase;letter-spacing:.3px}
td{padding:8px 8px 8px 0;border-bottom:1px solid #f5f5f5;cursor:pointer}
tr:last-child td{border-bottom:none}
tr:hover td{background:#f5f5f5}
.cat-dot{display:inline-block;width:10px;height:10px;border-radius:50%;margin-right:8px;vertical-align:middle}
.amt{text-align:right;font-variant-numeric:tabular-nums;white-space:nowrap}
.amt.neg{color:#e74c3c}
.amt.pos{color:#27ae60}
#txnSection{display:none;margin-top:24px;background:#fff;border-radius:10px;padding:20px;box-shadow:0 1px 3px rgba(0,0,0,.08)}
#txnSection h2{font-size:16px;margin-bottom:12px}
#txnSection h2 .close{float:right;cursor:pointer;color:#aaa;font-size:22px;line-height:1}
#txnSection h2 .close:hover{color:#333}
#txnTable{width:100%}
#txnTable td{border-bottom:1px solid #eee;padding:6px 8px 6px 0;cursor:default}
#txnTable th:nth-child(1){width:140px}
#txnTable th:nth-child(2){width:130px;text-align:right}
.cat-toggle{margin:0;width:16px;height:16px;cursor:pointer;vertical-align:middle}
#catTableContainer tr.dim{opacity:.4}
#catTableContainer tr.dim:hover td{background:transparent}
.hide-source #txnTable th:nth-child(4),.hide-source #txnTable td:nth-child(4){display:none}
.toggle-source{font-size:13px;margin-bottom:12px;display:inline-block;cursor:pointer;user-select:none}
.toggle-source input{vertical-align:middle;margin-right:4px}
</style>
</head>
<body>
<h1>&#x1F4CA; $REPORT_TITLE</h1>
<div class="summary" id="summaryCards"></div>
<div class="tabs" id="tabs">
    <div class="tab active" data-mode="expense">Расходы</div>
    <div class="tab" data-mode="income">Доходы</div>
    <div class="tab" data-mode="all">Всё</div>
</div>
<div class="chart-row">
    <div class="chart-box"><canvas id="pieChart"></canvas></div>
    <div class="cat-table" id="catTableContainer"></div>
</div>
<div id="txnSection">
    <h2><span id="txnTitle"></span><span class="close" onclick="closeTxn()">&times;</span></h2>
    <label class="toggle-source"><input type="checkbox" id="showSource" onchange="toggleSource()"> Показывать источник</label>
    <div id="txnWrapper">
    <table id="txnTable"><thead><tr><th onclick="sortBy('d')" style="cursor:pointer">Дата <span id="sortInd"></span></th><th>Сумма</th><th>Описание</th><th>Источник</th></tr></thead><tbody id="txnBody"></tbody></table>
    </div>
</div>
<script>
var DATA = $JSON_DATA ;
var hiddenCategories = {};

var COLORS = [
    '#e74c3c','#3498db','#2ecc71','#f39c12','#9b59b6','#1abc9c','#e67e22','#34495e',
    '#c0392b','#2980b9','#27ae60','#d35400','#8e44ad','#16a085','#f1c40f','#2c3e50',
    '#e88','#698','#8a9','#ba8','#9af','#fa8','#a78','#89a'
];

var currentMode = 'expense';
var chart = null;
var categoryColors = {};
var sortCol = 'd';
var sortDir = -1;

(function() {
    var i = 0;
    for (var name in DATA.categories) {
        categoryColors[name] = COLORS[i % COLORS.length];
        i++;
    }
})();

function filterTransactions(cat, mode) {
    if (mode === 'all') return cat;
    var txns = cat.transactions.filter(function(t) {
        var amt = parseFloat(t.a);
        return mode === 'expense' ? amt < 0 : amt > 0;
    });
    var total = txns.reduce(function(s, t) { return s + parseFloat(t.a); }, 0);
    return { total: total, count: txns.length, transactions: txns };
}

function filterData(mode) {
    var cats = {};
    for (var name in DATA.categories) {
        if (hiddenCategories[name]) continue;
        var d = DATA.categories[name];
        if (mode === 'all') {
            cats[name] = d;
        } else {
            var filtered = filterTransactions(d, mode);
            if (filtered.count === 0) continue;
            cats[name] = filtered;
        }
    }
    return cats;
}

function fmt(v) {
    var neg = v < 0;
    var s = Math.abs(v).toFixed(2);
    s = s.replace(/\\B(?=(\\d{3})+(?!\\d))/g, ' ');
    return (neg ? '\\u2212' : '') + s;
}

function fmtDate(s) {
    var p = s.split(' ');
    var d = p[0].split('-');
    return d[2] + '.' + d[1] + '.' + d[0] + ' ' + (p[1] || '');
}

function renderChart(mode) {
    var cats = filterData(mode);
    var labels = [];
    var values = [];
    var colors = [];
    for (var name in cats) {
        labels.push(name);
        values.push(Math.abs(cats[name].total));
        colors.push(categoryColors[name] || '#ccc');
    }
    if (chart) chart.destroy();
    if (labels.length > 0) {
        var ctx = document.getElementById('pieChart').getContext('2d');
        chart = new Chart(ctx, {
            type: 'pie',
            data: { labels: labels, datasets: [{ data: values, backgroundColor: colors, borderWidth: 0 }] },
            options: {
                responsive: true,
                plugins: {
                    legend: { display: false },
                    tooltip: {
                        callbacks: {
                            label: function(ctx) {
                                var total = ctx.dataset.data.reduce(function(a,b){return a+b}, 0);
                                var pct = ((ctx.parsed / total) * 100).toFixed(1);
                                return ' ' + ctx.label + ': ' + fmt(cats[ctx.label].total) + ' \u20bd (' + pct + '%)';
                            }
                        }
                    }
                },
                onClick: function(e, elements) {
                    if (elements.length > 0) {
                        showTransactions(this.data.labels[elements[0].index]);
                    }
                }
            }
        });
    }
    renderCategoryTable(mode, colors);
    renderSummary(cats);
}

function toggleCategory(name, el) {
    if (el.checked) {
        delete hiddenCategories[name];
    } else {
        hiddenCategories[name] = true;
    }
    renderChart(currentMode);
}

function renderSummary(cats) {
    var count = 0, income = 0, expense = 0;
    for (var name in cats) {
        var d = cats[name];
        count += d.count;
        if (d.total > 0) income += d.total;
        else expense += d.total;
    }
    document.getElementById('summaryCards').innerHTML =
        '<div class="card"><h3>Транзакции</h3><div class="val">' + count + '</div></div>'
        + '<div class="card"><h3>Доходы</h3><div class="val pos">' + fmt(income) + '</div></div>'
        + '<div class="card"><h3>Расходы</h3><div class="val neg">' + fmt(expense) + '</div></div>'
        + '<div class="card"><h3>Нетто</h3><div class="val ' + (income + expense < 0 ? 'neg' : 'pos') + '">' + fmt(income + expense) + '</div></div>';
}

function renderCategoryTable(mode, colors) {
    var allCats = {};
    for (var name in DATA.categories) {
        var d = DATA.categories[name];
        if (mode === 'all') {
            allCats[name] = d;
        } else {
            var filtered = filterTransactions(d, mode);
            if (filtered.count === 0) continue;
            allCats[name] = filtered;
        }
    }
    var html = '<table><thead><tr><th></th><th>Категория</th><th class="amt">Сумма</th><th class="amt">Кол-во</th></tr></thead><tbody>';
    var keys = Object.keys(allCats);
    for (var i = 0; i < keys.length; i++) {
        var name = keys[i];
        var d = allCats[name];
        var hidden = !!hiddenCategories[name];
        var color = categoryColors[name] || '#ccc';
        var rowClass = hidden ? ' class="dim"' : '';
        var rowClick = hidden ? '' : ' onclick="showTransactions(\\'' + name.replace(/'/g, "\\\\'") + '\\')"';
        html += '<tr' + rowClass + rowClick + '>'
            + '<td><input type="checkbox" class="cat-toggle" ' + (hidden ? '' : 'checked') + ' onchange="event.stopPropagation();toggleCategory(\\'' + name.replace(/'/g, "\\\\'") + '\\', this)"></td>'
            + '<td><span class="cat-dot" style="background:' + color + '"></span>' + name + '</td>'
            + '<td class="amt ' + (d.total < 0 ? 'neg' : 'pos') + '">' + fmt(d.total) + '</td>'
            + '<td class="amt">' + d.count + '</td></tr>';
    }
    html += '</tbody></table>';
    document.getElementById('catTableContainer').innerHTML = html;
}

function doSort(txns) {
    txns.sort(function(a, b) {
        var va = a[sortCol], vb = b[sortCol];
        if (sortCol === 'd') return (va < vb ? -1 : va > vb ? 1 : 0) * sortDir;
        if (sortCol === 'a') return (parseFloat(va) - parseFloat(vb)) * sortDir;
        return (va < vb ? -1 : va > vb ? 1 : 0) * sortDir;
    });
}

function sortBy(col) {
    if (col === sortCol) sortDir *= -1;
    else { sortCol = col; sortDir = -1; }
    var cur = document.getElementById('txnTitle').textContent;
    if (cur) showTransactions(cur);
}

function showTransactions(catName) {
    var d = DATA.categories[catName];
    if (!d) return;
    document.getElementById('txnTitle').textContent = catName;
    var sorted = (currentMode === 'all' ? d.transactions.slice() : filterTransactions(d, currentMode).transactions);
    doSort(sorted);
    var html = '';
    for (var i = 0; i < sorted.length; i++) {
        var t = sorted[i];
        var amt = parseFloat(t.a);
        var cls = amt < 0 ? 'neg' : 'pos';
        html += '<tr><td>' + fmtDate(t.d) + '</td><td class="amt ' + cls + '">' + fmt(amt) + '</td><td>' + t.n.slice(0,120) + '</td><td>' + t.s + '</td></tr>';
    }
    document.getElementById('txnBody').innerHTML = html;
    document.getElementById('sortInd').textContent = sortCol === 'd' ? (sortDir < 0 ? ' \u2193' : ' \u2191') : '';
    document.getElementById('txnSection').style.display = 'block';
}

function closeTxn() {
    document.getElementById('txnSection').style.display = 'none';
}

function toggleSource() {
    var el = document.getElementById('txnWrapper');
    if (document.getElementById('showSource').checked) {
        el.classList.remove('hide-source');
    } else {
        el.classList.add('hide-source');
    }
}
toggleSource();

(function() {
    var tabs = document.querySelectorAll('.tab');
    for (var i = 0; i < tabs.length; i++) {
        tabs[i].addEventListener('click', function() {
            var active = document.querySelectorAll('.tab');
            for (var j = 0; j < active.length; j++) active[j].classList.remove('active');
            this.classList.add('active');
            currentMode = this.dataset.mode;
            closeTxn();
            renderChart(currentMode);
        });
    }
})();

renderChart('expense');
</script>
</body>
</html>"""


def find_latest_dir():
    dirs = [d for d in os.listdir(RESULTS_DIR)
            if os.path.isdir(os.path.join(RESULTS_DIR, d)) and re.match(r"\d{4}-\d{2}-\d{2}_\d+", d)]
    if not dirs:
        return None
    dirs.sort(key=lambda d: int(d.split("_")[-1]))
    return os.path.join(RESULTS_DIR, dirs[-1])


def main():
    date_from = None
    date_to = None
    mode = None
    args = sys.argv[1:]
    for i, a in enumerate(args):
        if a == "--year" and i + 1 < len(args):
            y = int(args[i + 1])
            date_from = datetime(y, 1, 1)
            date_to = datetime(y, 12, 31, 23, 59)
            mode = ("year", y)
        elif a == "--month" and i + 1 < len(args):
            parts = args[i + 1].split("-")
            y, m = int(parts[0]), int(parts[1])
            date_from = datetime(y, m, 1)
            if m == 12:
                date_to = datetime(y + 1, 1, 1)
            else:
                date_to = datetime(y, m + 1, 1)
            mode = ("month", y, m)
        elif a == "--from" and i + 1 < len(args):
            date_from = datetime.strptime(args[i + 1], "%Y-%m-%d")
            mode = "range" if mode is None else mode
        elif a == "--to" and i + 1 < len(args):
            date_to = datetime.strptime(args[i + 1] + " 23:59:59", "%Y-%m-%d %H:%M:%S")
            mode = "range" if mode is None else mode

    latest_dir = find_latest_dir()
    if not latest_dir:
        print("No results directories found")
        return

    xlsx_path = os.path.join(latest_dir, "excel.xlsx")
    if not os.path.exists(xlsx_path):
        print(f"excel.xlsx not found in {latest_dir}")
        return

    wb = load_workbook(xlsx_path)
    ws = wb["Сводный"]

    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {h: i for i, h in enumerate(headers)}

    totals = {"count": 0, "income": 0.0, "expense": 0.0}
    categories = defaultdict(lambda: {"total": 0.0, "count": 0, "transactions": []})

    for row in ws.iter_rows(min_row=2, values_only=True):
        dt = row[col["Дата"]]
        if isinstance(dt, datetime):
            if date_from and dt < date_from:
                continue
            if date_to and dt > date_to:
                continue
            date_str = dt.strftime("%Y-%m-%d %H:%M")
        else:
            date_str = str(dt or "")

        cat = str(row[col["Категория"]]) if row[col["Категория"]] else "Неизвестно"
        raw = row[col["Сумма"]]
        amount = float(raw) if raw is not None else 0.0

        txn = {
            "d": date_str,
            "a": round(amount, 2),
            "n": str(row[col["Описание"]] or ""),
            "s": str(row[col["Источник"]] or "")
        }
        categories[cat]["total"] += amount
        categories[cat]["count"] += 1
        categories[cat]["transactions"].append(txn)

        if amount > 0:
            totals["income"] += amount
        else:
            totals["expense"] += amount
        totals["count"] += 1

    data = {
        "summary": {
            "count": totals["count"],
            "income": round(totals["income"], 2),
            "expense": round(totals["expense"], 2),
            "net": round(totals["income"] + totals["expense"], 2)
        },
        "categories": {}
    }

    for cat_name in sorted(categories, key=lambda c: -abs(categories[c]["total"])):
        cd = categories[cat_name]
        cd["total"] = round(cd["total"], 2)
        cd["transactions"].sort(key=lambda t: t["d"])
        data["categories"][cat_name] = cd

    json_data = json.dumps(data, ensure_ascii=False, indent=2)

    if mode and mode[0] == "year":
        y = mode[1]
        report_title = f"Финансовый отчёт за {y} год"
        report_name = f"report-{y}.html"
    elif mode and mode[0] == "month":
        y, m = mode[1], mode[2]
        report_title = f"Финансовый отчёт за {MONTHS_RU[m]} {y}"
        report_name = f"report-{y}-{m:02d}.html"
    elif mode == "range":
        fstr = date_from.strftime("%d.%m.%Y") if date_from else "начала"
        tstr = date_to.strftime("%d.%m.%Y") if date_to else "конца"
        report_title = f"Финансовый отчёт: {fstr} — {tstr}"
        fpart = date_from.strftime("%Y-%m-%d") if date_from else "0000-00-00"
        tpart = date_to.strftime("%Y-%m-%d") if date_to else "9999-99-99"
        report_name = f"report-{fpart}_{tpart}.html"
    else:
        report_title = "Финансовый отчёт"
        report_name = "report.html"

    html = HTML_TPL.replace("$JSON_DATA", json_data).replace("$REPORT_TITLE", report_title)

    report_path = os.path.join(latest_dir, report_name)
    with open(report_path, "w", encoding="utf-8") as f:
        f.write(html)

    print(f"Report saved: {report_path}")


if __name__ == "__main__":
    main()
